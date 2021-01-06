'''
分为三个等级：
level_1 :   会执行python代码

level_2 :   会编写功能性函数
	1、从data文件夹中读取excel数据
	2、将data文件夹中的数据进行分组统计并写入到result文件夹的excel文件中
	3、将程序整理后的数据写入result文件夹的excel文件中的指定位置
	4、将result文件夹中的excel文件通过邮件的形式发出来

level_3 :   会整合功能性函数，实现一个可用的程序
	整合level_2中的功能函数，执行主程序的时候能够自动将结果通过邮件的形式发出来
'''


import pandas as pd
import openpyxl
from tool.send_email import report_send_email

#加载数据
def load_data(filename = 'data/data.xlsx'):
    df = pd.read_excel(filename)
    return df

# 将数据进行分组统计
def data_process(df):
    df_group = df.groupby('部门', as_index=False).sum()
    return df_group

# 写入到result文件夹的excel文件中
def save_data(df,filename='result/result.xlsx'):
    df.to_excel(filename,index=False)



def save_data_excel(df_group):

    wb = openpyxl.load_workbook('result/result_1.xlsx')# 打开服务器存储路径下的excel文件

    #读取sheet表
    sheet = wb['Sheet1']    #Sheet1表

    #清除历史数据
    for i in range(2,sheet.max_row + 1):
        for j in range(1,sheet.max_column +1):
            sheet.cell(row = i,column = j).value = ''

    row_num = df_group.shape[0]
    col_num = df_group.shape[1]
    #填充结果数据
    for i in range(2,row_num+2):
        for j in range(1,col_num+1):
            sheet.cell(row=i,column=j).value = df_group.iloc[i-2][j-1]
    wb.save('result/result_2.xlsx')   #保存excel文件



if __name__ == '__main__':
    df = load_data()
    print(df)
    df_group = data_process(df)
    print(df_group)
    # 将整理后的数据写入excel的指定位置
    save_data_excel(df_group)
    # 将result_2.xlsx通过邮箱发送到'576591043@qq.com'
    # (from_mail_pass已经屏蔽，可以自己设置）教程见网址https://www.cnblogs.com/xiaodai12138/p/10483158.html
    report_send_email('C:/Users/KZF/Desktop/test1/result/result_2.xlsx','1369157509@qq.com','15172054680@163.com','BTHSVBTDJRABAUOJ')






